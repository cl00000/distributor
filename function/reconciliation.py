# function/reconciliation.py
import sys
from io import StringIO


class OutputRedirector:
    """重定向输出到GUI的类"""

    def __init__(self, callback):
        self.callback = callback
        self.buffer = ""

    def write(self, message):
        self.buffer += message
        # 按行处理，每遇到换行就发送一次
        if '\n' in self.buffer:
            lines = self.buffer.split('\n')
            for line in lines[:-1]:  # 发送完整行
                if line.strip():
                    self.callback(line)
            self.buffer = lines[-1]  # 保留未完成的半行

    def flush(self):
        # 发送缓冲区中剩余的内容
        if self.buffer:
            self.callback(self.buffer)
            self.buffer = ""


def run_reconciliation_with_gui(output_callback=None):
    """
    在GUI环境中运行对账功能

    Args:
        output_callback: 回调函数，用于将输出发送到GUI界面
    """
    # 保存原始标准输出
    old_stdout = sys.stdout

    try:
        if output_callback:
            # 创建重定向器
            redirector = OutputRedirector(output_callback)
            sys.stdout = redirector

        # 调用原有的处理逻辑
        return process_all_files()

    except Exception as e:
        if output_callback:
            output_callback(f"❌ 处理失败: {str(e)}")
        return False
    finally:
        # 恢复原始标准输出
        sys.stdout = old_stdout


def process_all_files():
    """原有的处理逻辑，包装成函数"""
    # ===============================
    # 路径配置 - 已根据要求修改
    # ===============================
    import os
    data_folder = r"D:\分销对账"
    mapping_folder = r"D:\分销对账\编码表"
    mapping_file = os.path.join(mapping_folder, "编码.xlsx")

    # 检查路径是否存在
    if not os.path.exists(data_folder):
        raise FileNotFoundError(f"数据文件夹不存在: {data_folder}")

    if not os.path.exists(mapping_file):
        raise FileNotFoundError(f"编码文件不存在: {mapping_file}")

    print(f"数据文件夹: {data_folder}")
    print(f"编码文件: {mapping_file}")

    # ===============================
    # 读取编码表并建立映射关系
    # ===============================
    import pandas as pd
    map_df = pd.read_excel(mapping_file)

    # 建立含税分销商映射
    tax_distributor_map = {}
    for _, row in map_df.iterrows():
        tax_distributor = row.get("含税分销商")
        if pd.notna(tax_distributor):
            # 清理分销商名称，去掉空格
            tax_distributor_clean = str(tax_distributor).strip()
            tax_distributor_map[tax_distributor_clean] = True

    print(f"含税分销商列表: {list(tax_distributor_map.keys())}")

    # 建立编码信息字典
    code_info = {}
    for _, row in map_df.iterrows():
        code = str(row["货品商家编码"])
        code_info[code] = {
            "name": str(row["名称"]),
            "type": str(row["产品类型"]),
            "price": float(row["供货价"]),  # 标准供货价
            "tax_price": float(row["供货价（含税）"]) if pd.notna(row["供货价（含税）"]) else None  # 含税供货价
        }

    # ===============================
    # 样式
    # ===============================
    from openpyxl.styles import Border, Side, Font, Alignment
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    red_font = Font(color="FF0000")

    success_count = 0
    error_count = 0
    multiple_code_files = []  # 记录有多重编码字段的文件

    # ===============================
    # 处理文件
    # ===============================
    import glob
    # 获取所有Excel文件（支持.xls和.xlsx）
    excel_files = glob.glob(os.path.join(data_folder, "*.xls")) + glob.glob(os.path.join(data_folder, "*.xlsx"))

    if not excel_files:
        print(f"❌ 在文件夹 {data_folder} 中未找到Excel文件")
        return False

    for file_path in excel_files:
        try:
            file_name = os.path.basename(file_path)
            print(f"正在处理: {file_name}")

            # ===============================
            # 解析文件名，确定使用哪种价格
            # ===============================
            file_stem = os.path.splitext(file_name)[0]  # 例如 "36号-上海帝亚"

            # 提取"-"前面的分销商编号
            if "-" in file_stem:
                distributor_code = file_stem.split("-")[0].strip()
            else:
                distributor_code = file_stem.strip()

            print(f"  分销商编号: {distributor_code}")

            # 判断是否使用含税价格
            use_tax_price = distributor_code in tax_distributor_map

            if use_tax_price:
                print(f"  ✓ 使用含税价格（供货价（含税））")
            else:
                print(f"  ✓ 使用标准价格（供货价）")

            # ===============================
            # 使用 openpyxl 直接读取 Excel 文件检查表头
            # ===============================
            from openpyxl import load_workbook

            # 先检查文件表头
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb["Sheet1"]

            # 获取第一行所有单元格的值
            header_values = []
            for cell in ws[1]:
                header_values.append(cell.value)

            # 统计"商家编码"出现的次数
            merchant_code_count = 0
            merchant_code_positions = []
            for col_idx, value in enumerate(header_values, 1):
                if value == "商家编码":
                    merchant_code_count += 1
                    merchant_code_positions.append(col_idx)

            # ===============================
            # 如果有多个"商家编码"字段，跳过处理
            # ===============================
            if merchant_code_count > 1:
                print(f"❌ 跳过 {file_name}：发现 {merchant_code_count} 个'商家编码'字段")
                print(f"   位置：第 {', '.join(map(str, merchant_code_positions))} 列")
                print(f"   请检查Excel文件，删除多余的'商家编码'列")

                # 记录这个文件
                multiple_code_files.append(file_name)
                error_count += 1

                # 关闭只读工作簿
                wb.close()
                continue

            # 关闭只读工作簿
            wb.close()

            # ===============================
            # 如果没有"商家编码"字段，也跳过
            # ===============================
            if merchant_code_count == 0:
                print(f"跳过 {file_name}：未找到'商家编码'列")
                print(f"   可用列名：{header_values}")
                error_count += 1
                continue

            # ===============================
            # 使用 pandas 读取数据
            # ===============================
            df = pd.read_excel(file_path, sheet_name="Sheet1")

            # 再次确认只有一个"商家编码"列
            merchant_code_cols = [col for col in df.columns if str(col).strip() == "商家编码"]
            if len(merchant_code_cols) > 1:
                print(f"❌ 跳过 {file_name}：pandas检测到 {len(merchant_code_cols)} 个'商家编码'列")
                print(f"   列名：{merchant_code_cols}")
                error_count += 1
                continue

            from collections import defaultdict
            import re
            code_counter = defaultdict(int)
            unmatched_codes = set()
            missing_price_names = set()

            # ===============================
            # 解析商家编码
            # ===============================
            for cell in df["商家编码"].dropna():
                items = str(cell).split(";")
                normal_total = 0
                gift_items = []

                for item in items:
                    m = re.match(r"(.+?)(?:\*(\d+))?$", item.strip())
                    if not m:
                        continue

                    code = m.group(1)
                    qty = int(m.group(2)) if m.group(2) else 1

                    info = code_info.get(code)
                    if not info:
                        unmatched_codes.add(code)
                        continue

                    if info["type"] == "赠品":
                        gift_items.append((code, qty))
                    else:
                        normal_total += qty
                        code_counter[code] += qty

                gift_total = sum(q for _, q in gift_items)

                if normal_total == 0:
                    for code, qty in gift_items:
                        code_counter[code] += qty
                    continue

                extra = gift_total - normal_total
                if extra > 0:
                    for code, qty in gift_items:
                        use = min(qty, extra)
                        code_counter[code] += use
                        extra -= use
                        if extra <= 0:
                            break

            # ===============================
            # 汇总到【名称】并根据分销商选择价格
            # ===============================
            final = {}
            for code, qty in code_counter.items():
                info = code_info[code]
                name = info["name"]

                # 根据是否使用含税价格选择价格
                if use_tax_price and info["tax_price"] is not None:
                    price = info["tax_price"]
                else:
                    price = info["price"]

                # 供货价缺失判断
                if pd.isna(price) or price == 0:
                    missing_price_names.add(name)

                if name not in final:
                    final[name] = {
                        "数量": 0,
                        "供货价": price if not pd.isna(price) else ""
                    }

                final[name]["数量"] += qty

            # ===============================
            # 打开 Excel 进行写入
            # ===============================
            wb = load_workbook(file_path)
            ws = wb["Sheet1"]

            # 找「商家编码」列
            code_col = None
            for c in range(1, ws.max_column + 1):
                if ws.cell(1, c).value == "商家编码":
                    code_col = c
                    break
            if not code_col:
                print(f"跳过 {file_name}：未找到'商家编码'列")
                continue

            start_col = code_col + 4  # 间隔 3 列

            # ===============================
            # 解除旧合并（关键）
            # ===============================
            for rng in list(ws.merged_cells.ranges):
                if rng.min_col >= start_col:
                    ws.unmerge_cells(str(rng))

            # ===============================
            # 清空旧结果区
            # ===============================
            for r in range(1, ws.max_row + 1):
                for c in range(start_col, ws.max_column + 1):
                    ws.cell(r, c).value = None
                    ws.cell(r, c).border = Border()

            # ===============================
            # 表头（保持"供货价"不变）
            # ===============================
            headers = ["分销商", "名称", "供货价", "数量", "售后处理费", "金额"]
            for i, h in enumerate(headers):
                cell = ws.cell(1, start_col + i, h)
                cell.border = border

            # ===============================
            # 列字母（一次算好）
            # ===============================
            from openpyxl.utils import get_column_letter
            price_col = get_column_letter(start_col + 2)
            qty_col = get_column_letter(start_col + 3)
            fee_col = get_column_letter(start_col + 4)
            amt_col = get_column_letter(start_col + 5)

            # ===============================
            # 写数据
            # ===============================
            start_row = 2
            r = start_row

            for name, info in final.items():
                ws.cell(r, start_col + 1, name)
                ws.cell(r, start_col + 2, info["供货价"])
                ws.cell(r, start_col + 3, -info["数量"])

                ws.cell(r, start_col + 4, f"={qty_col}{r}*1")
                ws.cell(
                    r,
                    start_col + 5,
                    f"={price_col}{r}*{qty_col}{r}-{fee_col}{r}"
                )
                r += 1

            end_row = r - 1

            # ===============================
            # 分销商合并（安全）
            # ===============================
            if end_row >= start_row:
                ws.cell(start_row, start_col).value = file_stem
                ws.merge_cells(
                    start_row=start_row,
                    start_column=start_col,
                    end_row=end_row,
                    end_column=start_col
                )
                # 添加垂直水平居中样式（修复弃用警告）
                ws.cell(start_row, start_col).alignment = Alignment(
                    horizontal='center',
                    vertical='center'
                )

            # ===============================
            # 汇总行
            # ===============================
            total_row = end_row + 1
            ws.cell(total_row, start_col + 1, "合计")
            ws.cell(
                total_row,
                start_col + 3,
                f"=SUM({qty_col}{start_row}:{qty_col}{end_row})"
            )
            ws.cell(
                total_row,
                start_col + 4,
                f"=SUM({fee_col}{start_row}:{fee_col}{end_row})"
            )
            ws.cell(
                total_row,
                start_col + 5,
                f"=SUM({amt_col}{start_row}:{amt_col}{end_row})"
            )

            # ===============================
            # 边框和居中（表头 + 数据 + 合计）
            # ===============================
            for row in range(1, total_row + 1):
                for col in range(start_col, start_col + len(headers)):
                    cell = ws.cell(row, col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # ===============================
            # 列宽
            # ===============================
            ws.column_dimensions[get_column_letter(start_col)].width = 22  # 分销商
            ws.column_dimensions[get_column_letter(start_col + 1)].width = 22  # 名称
            ws.column_dimensions[get_column_letter(start_col + 2)].width = 15  # 供货价
            ws.column_dimensions[get_column_letter(start_col + 3)].width = 15
            ws.column_dimensions[get_column_letter(start_col + 4)].width = 15
            ws.column_dimensions[get_column_letter(start_col + 5)].width = 15

            for i in range(1, 4):  # 间隔列
                ws.column_dimensions[get_column_letter(code_col + i)].width = 6

            # ===============================
            # 未匹配编码提示（不影响列宽）
            # ===============================
            warn_row = total_row + 2

            # 未匹配编码
            if unmatched_codes:
                ws.cell(
                    warn_row,
                    start_col,
                    "⚠ 以下商家编码未在编码表中匹配，请人工核对"
                ).font = red_font
                ws.cell(
                    warn_row + 1,
                    start_col,
                    ", ".join(sorted(unmatched_codes))
                ).font = red_font
                warn_row += 3

            # 缺失供货价
            if missing_price_names:
                ws.cell(
                    warn_row,
                    start_col,
                    "⚠ 以下商品未配置供货价，请补充后重新计算"
                ).font = red_font
                ws.cell(
                    warn_row + 1,
                    start_col,
                    ", ".join(sorted(missing_price_names))
                ).font = red_font

            # ===============================
            # 价格类型提示
            # ===============================
            if use_tax_price:
                ws.cell(
                    warn_row + 2 if warn_row > total_row + 2 else total_row + 2,
                    start_col,
                    f"📝 注：本表使用含税价格（供货价（含税））"
                )

            wb.save(file_path)
            print(f"✅ 已处理：{file_name}")
            success_count += 1

        except Exception as e:
            print(f"❌ 处理失败：{file_name} → {e}")
            import traceback
            traceback.print_exc()
            error_count += 1

    # ===============================
    # 输出汇总信息
    # ===============================
    print(f"\n处理完成！成功：{success_count} 个文件，失败：{error_count} 个文件")

    if multiple_code_files:
        print(f"\n⚠ 以下文件因有多个'商家编码'字段未处理：")
        for file_name in multiple_code_files:
            print(f"  • {file_name}")
        print(f"请检查这些文件，删除多余的'商家编码'列后重新执行")

    return True