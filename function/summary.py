def run_summary(output_callback=None):
    import os
    from datetime import datetime
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Border, Side, Alignment, Font

    def log(msg):
        if output_callback:
            output_callback(msg)
        else:
            print(msg)

    # ===== 路径
    base_dir = r"D:\分销对账"
    summary_dir = os.path.join(base_dir, "汇总表")
    os.makedirs(summary_dir, exist_ok=True)

    # ===== 获取上个月（用于文件名）和前两个月（用于标题）
    today = datetime.today()
    year = today.year
    month = today.month - 1
    if month == 0:
        month = 12
        year -= 1

    # 获取前两个月（用于标题）
    title_month = today.month - 2
    title_year = today.year
    if title_month <= 0:
        title_month += 12
        title_year -= 1

    month_label = f"{month}月"
    title_month_label = f"{title_month}月"

    summary_file = os.path.join(
        summary_dir, f"{year}-{month_label}售后汇总.xlsx"
    )
    log(f"📂 生成汇总表：{summary_file}")

    wb = Workbook()
    ws = wb.active
    ws.title = "售后汇总"

    # 先定义 headers
    headers = [
        "分销商", "名称", "供货价", "数量",
        "售后处理费", "应返还金额",
        "售后处理费（合计）", "售后返还总额",
        f"{month_label}营业额", f"{month_label}有效营业额"
    ]

    # ===== 添加标题行（横向合并）
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(1, 1, f"{title_month_label}售后数据")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.font = Font(size=14, bold=True)

    # ===== 设置标题行高度
    ws.row_dimensions[1].height = 30

    # 添加表头到第二行
    ws.append(headers)

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ===== 设置表头样式（第二行）
    for col in range(1, len(headers) + 1):
        cell = ws.cell(2, col)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        cell.font = Font(bold=True)

    write_row = 3  # 从第三行开始写数据

    # ===== 收尾分销商
    def finalize_distributor(start_row, end_row):
        if start_row is None or end_row < start_row:
            return

        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=end_row, end_column=1)
        ws.cell(start_row, 1).alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(start_row=start_row, start_column=7,
                       end_row=end_row, end_column=7)
        ws.cell(start_row, 7, f"=SUM(E{start_row}:E{end_row})").alignment = Alignment(horizontal="center",
                                                                                      vertical="center")

        ws.merge_cells(start_row=start_row, start_column=8,
                       end_row=end_row, end_column=8)
        ws.cell(start_row, 8, f"=SUM(F{start_row}:F{end_row})").alignment = Alignment(horizontal="center",
                                                                                      vertical="center")

        ws.merge_cells(start_row=start_row, start_column=9,
                       end_row=end_row, end_column=9)
        ws.cell(start_row, 9).alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(start_row=start_row, start_column=10,
                       end_row=end_row, end_column=10)
        ws.cell(start_row, 10, f"=H{start_row}+I{start_row}").alignment = Alignment(horizontal="center",
                                                                                    vertical="center")

    # ===== 读取对账文件
    for file in os.listdir(base_dir):
        if not file.endswith((".xls", ".xlsx")) or file.startswith("~$"):
            continue

        wb_src = load_workbook(os.path.join(base_dir, file), data_only=False)
        ws_src = wb_src["Sheet1"]

        start_col = None
        for i, cell in enumerate(ws_src[1], 1):
            if cell.value == "分销商":
                start_col = i
                break

        if not start_col:
            wb_src.close()
            continue

        col = {
            "分销商": start_col,
            "名称": start_col + 1,
            "供货价": start_col + 2,
            "数量": start_col + 3,
        }

        current_distributor = None
        distributor_start_row = None
        last_distributor = None

        r = 2
        while r <= ws_src.max_row:
            name = ws_src.cell(r, col["名称"]).value
            if not name or str(name).strip() == "合计":
                r += 1
                continue

            raw = ws_src.cell(r, col["分销商"]).value
            distributor = raw if raw else last_distributor
            last_distributor = distributor

            if distributor != current_distributor:
                if current_distributor is not None:
                    finalize_distributor(distributor_start_row, write_row - 1)
                current_distributor = distributor
                distributor_start_row = write_row
                ws.cell(write_row, 1, distributor)

            ws.cell(write_row, 2, name)
            ws.cell(write_row, 3, ws_src.cell(r, col["供货价"]).value)
            ws.cell(write_row, 4, ws_src.cell(r, col["数量"]).value)
            ws.cell(write_row, 5, f"=D{write_row}*1")
            ws.cell(write_row, 6, f"=C{write_row}*D{write_row}-E{write_row}")

            # 设置当前行所有单元格的水平和垂直居中对齐和边框
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(write_row, col_num)
                # 根据列类型设置不同的水平对齐方式
                if col_num == 2:  # 名称列，左对齐
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif col_num in [3, 4, 5, 6]:  # 供货价、数量、售后处理费、应返还金额列，水平和垂直居中
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:  # 其他列，垂直居中
                    cell.alignment = Alignment(vertical="center")
                cell.border = border

            write_row += 1
            r += 1

        if current_distributor:
            finalize_distributor(distributor_start_row, write_row - 1)

        wb_src.close()

    # ===== 全表合计行
    total_row = write_row

    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
    ws.cell(total_row, 1, "合计").alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(total_row, 1).font = Font(bold=True)

    ws.cell(total_row, 7, f"=SUM(G3:G{total_row - 1})")
    ws.cell(total_row, 8, f"=SUM(H3:H{total_row - 1})")
    ws.cell(total_row, 9, f"=SUM(I3:I{total_row - 1})")
    ws.cell(total_row, 10, f"=SUM(J3:J{total_row - 1})")

    # 设置合计行样式
    for c in range(1, len(headers) + 1):
        cell = ws.cell(total_row, c)
        cell.border = border
        if c == 2:  # 名称列，左对齐
            cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c in [3, 4, 5, 6, 7, 8, 9, 10]:  # 数值列，水平和垂直居中
            cell.alignment = Alignment(horizontal="center", vertical="center")
        if c >= 7:  # 数值列加粗
            cell.font = Font(bold=True)

    # ===== 行高 & 边框（包括合计行）
    for r in range(2, total_row + 1):
        ws.row_dimensions[r].height = 22
        for c in range(1, len(headers) + 1):
            cell = ws.cell(r, c)
            # 确保所有单元格都有边框
            cell.border = border
            # 确保数值列（3-6列）水平和垂直居中对齐
            if c in [3, 4, 5, 6]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # 确保其他列至少垂直居中对齐
            elif cell.alignment is None:
                cell.alignment = Alignment(vertical="center")
            elif c != 2:  # 名称列保持左对齐
                # 保留原有的水平对齐，设置垂直居中
                cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                           vertical="center",
                                           wrap_text=cell.alignment.wrap_text)

    # ===== 列宽
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 18

    wb.save(summary_file)
    return True